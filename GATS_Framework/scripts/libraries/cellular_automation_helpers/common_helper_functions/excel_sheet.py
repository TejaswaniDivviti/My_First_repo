import os
import pandas as pd
import xlsxwriter
import zipfile
import itertools
from zipfile import ZipFile
from datetime import datetime

import automation_helpers.globalconstants as gc

def sheet_creation(workbook,sheet_name, headings):
    """ Creating and returning sheet for each message """
    worksheet = workbook.add_worksheet(sheet_name)
    worksheet.set_column('A:C', 70)
    worksheet.set_row(0, 30)

    worksheet.set_column('A:A', 13)
    worksheet.set_column('B:B', 17)
    worksheet.set_column('C:C', 17)
    worksheet.set_column('D:D', 17)
    worksheet.set_column('E:E', 30)
    worksheet.set_column('F:F', 30)
    bold = workbook.add_format({'bold': 1, 'align': 'right'})
    worksheet.write_row('A1', headings, bold)
    return worksheet
path = ""
def create_excel(log_type):
    global path
    path = gc.IMAGE_FOLDER+"/../../../Merged_Output/"
    if os.path.isdir(path) == False:
        os.makedirs(path)
    workbook = xlsxwriter.Workbook("{0}/Robot_Results.xlsx".format(path))

    headings = ['S.No', 'TestCaseId', 'IterationNo', "PASS/FAIL/NA", "Logcat Logs", "Modem Logs", "Reasons"]
    worksheet = sheet_creation(workbook, "Result", headings)
    Values1 = []
    center = workbook.add_format({'align': 'center'})
    Values1 = gc.EXCEL_DATA.values()
    exec_status = []
    android_log_path = []
    output_path = []
    msg = []
    for v in Values1:
        exec_status.append(v[0])
        android_log_path.append(v[1])
        output_path.append((v[1]+"/../"))
        msg.append(v[2])

    print(output_path)
    iter_no = []
    tc_name = []
    for key_name in list(gc.EXCEL_DATA.keys()):
        y = key_name.rsplit("_", 1) 
        it = str(y[1])
        iter_no.append(it)
        tc_name.append(y[0])

    row = 1
    col = 0
    for (a,b,c,d,e,f) in itertools.zip_longest(tc_name, iter_no, exec_status, android_log_path, output_path, msg):        
            if b != None: 
                worksheet.write(row , col, row, center)
                worksheet.write(row , col + 1, a, center)
                worksheet.write(row , col + 2, b, center)
                worksheet.write(row , col + 3, c, center)
                worksheet.write(row , col + 4, d, center)
                if int(log_type) == 0:
                    worksheet.write(row , col + 5, d, center)
                elif int(log_type) == 1:
                    worksheet.write(row , col + 5, e, center)
                worksheet.write(row , col + 6, f, center)

            row += 1
    
    workbook.close()
    createSmartSummary()

import openpyxl
import enum
from openpyxl.styles import PatternFill,Border,Side
from openpyxl.styles.alignment import Alignment
class Color():
    class colr(enum.Enum):
        FFBF00 = 1
        solid = 2
        PASS="PASS"
        FAIL="FAIL"
        NA="NA"
        Thin="thin"


    def __init__(self,filename):
        try:
            '''Given a path name'''
            self.path = filename

            self.wb = openpyxl.load_workbook(self.path)
            '''Creating a new sheet in same excel'''
            self.ws = self.wb.create_sheet("smart summary",2)
        except Exception as e:
            print("no excel sheet", e)

    def read(self):
            '''Activating sheet'''
            self.sh=self.wb.active

            '''accessing a B and D column'''
            want_data = self.sh['B']
            want_data1 = self.sh['D']
            list2 = []
            list1 = []
            TC = 0
            '''Zip is used to compare B and D col at a time'''
            for k, v in zip(want_data, want_data1):
                '''i gives total Testcase id'''
                i = k.value
                '''j gives TC's result whether passed, failed, none'''
                j = v.value
                list2.append(i)
                list1.append(j)
                '''Assigning values Pass,fail, none'''
                PASS = list1.count(self.colr.PASS.value)
                FAIL = list1.count(self.colr.FAIL.value)
                NA = list1.count(self.colr.NA.value)
                '''Incrementing each TC one after the other'''
                TC += 1
            return TC,PASS,FAIL, NA

    def style(self):
        '''Color and border'''
        #coloring a row1
        for row in self.ws.iter_rows(min_row=1,max_col=6):
            for cell in row:
                    '''Coloring a row'''
                    cell.fill=PatternFill(start_color=self.colr(1).name,fill_type=self.colr(2).name)
                    '''bordering to a cell'''
                    thin_border=Border(left=Side(style=self.colr.Thin.value),right=Side(style=self.colr.Thin.value),top=Side(style=self.colr.Thin.value),bottom=Side(style=self.colr.Thin.value))
        for i in range(1,3):
                for j in range(1,7):
                    self.ws.cell(row=i,column=j).border=thin_border

        '''merging two cell'''
        self.ws. merge_cells('A1:A2')


        '''keeping values at center in row 2'''
        col_rows=["B2","D2","E2","F2"]
        for k in col_rows:
                self.ws[k].alignment = Alignment(horizontal="center")


        '''naming the row1'''
        header=["Complete summary","No of TC's", "Deviceused","Passed","Failed","None"]
        col_val = 1
        for r in header:
                self.ws.cell(row=1,column= col_val).value = r
                col_val+=1

    def write(self):
            '''appending values to a sheet'''
            TC,PASS,FAIL,NA = self.read()
            TC = int(TC) -1
            row=[TC,("LAVA,MI10T,PIXEL 3A"),PASS,FAIL,NA]
            col=2
            for x in row:
                self.ws.cell(row=2,column=col).value=x
                col+=1


    def save(self):
        '''saving a excel sheet'''
        self.wb.save(self.path)

def createSmartSummary():
    s= Color(filename="{0}/Robot_Results.xlsx".format(path))
    s.read()
    s.style()
    s.write()
    s.save()
